import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Diretório onde os arquivos de saída estão localizados
output_dir = 'output'

# Nome do arquivo de resposta
response_filename = 'ipc_summary.xlsx'

# Função para processar cada arquivo de saída
def process_ipc_stats_file(filename):
    section_name = filename.split('_ipc_stats.txt')[0]
    ipc_stats = []

    with open(os.path.join(output_dir, filename), 'r') as file:
        # next(file)  # pular o cabeçalho
        for line in file:
            ipc_type, count, percentage = line.strip().split()
            ipc_stats.append((ipc_type, float(percentage)))

    # Ordenar IPCs por porcentagem em ordem decrescente
    ipc_stats.sort(key=lambda x: x[1], reverse=True)
    return section_name, ipc_stats

# Lista de arquivos de saída
ipc_files = [f for f in os.listdir(output_dir) if f.endswith('_ipc_stats.txt')]

# Criar um novo workbook
wb = Workbook()
ws = wb.active
ws.title = "IPC Summary"

# Dicionário para armazenar cores
ipc_colors = {}

# Função para gerar uma cor hexadecimal
def generate_color(value):
    return PatternFill(start_color=value, end_color=value, fill_type='solid')

# Lista de cores para serem utilizadas
colors = ["FFB6C1", "ADD8E6", "90EE90", "FFD700", "FF69B4", "FFA500", "DA70D6", "32CD32"]

# Processar todos os arquivos e adicionar dados na planilha
row = 1
for ipc_file in ipc_files:
    section_name, ipc_stats = process_ipc_stats_file(ipc_file)
    ws.cell(row=row, column=1, value=section_name)
    
    col = 2
    for ipc, percentage in ipc_stats:
        if ipc not in ipc_colors:
            ipc_colors[ipc] = colors[len(ipc_colors) % len(colors)]
        cell = ws.cell(row=row, column=col, value=f"{ipc}({percentage:.2f})")
        cell.fill = generate_color(ipc_colors[ipc])
        col += 1

    row += 1

# Ajustar largura das colunas
for col in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col)].width = 20

# Salvar o arquivo Excel
wb.save(response_filename)

print(f"Arquivo de resumo criado: {response_filename}")
